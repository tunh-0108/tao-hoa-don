"""
export_excel.py
---------------
Tạo file Excel OUTPUT để import vào VinInvoice.

File output có đúng 41 cột (theo file mẫu MauUploadHDMTT). Phần lớn cột để trống,
chỉ điền những cột có logic mapping. Mỗi giá trị được ghi dạng CHUỖI (text) để
giống hệt file mẫu (tránh VinInvoice import bị lệch kiểu dữ liệu).

Quy tắc gộp dòng (quan trọng):
  - Mỗi DÒNG trong bảng hóa đơn = 1 hóa đơn = 1 MaHD (đánh số tăng dần từ 1).
  - Trong 1 dòng, mỗi NHÓM hàng hóa có dữ liệu sẽ tạo ra 1 DÒNG output (cùng MaHD).
  - Các dòng output cùng MaHD chỉ khác nhau ở: TenHangHoa, DonViTinh, SoLuong, DonGia, ThanhTien.
"""

import io
import openpyxl
from openpyxl.styles import Font

import invoice_logic as L

# 41 tên cột của file output, đúng thứ tự A -> AO
OUTPUT_HEADERS = [
    "MaHD", "LoaiHoaDon", "HoaDonLienQuanNgoaiHeThong", "MauSoHoaDonLienQuan",
    "KyHieuHoaDonLienQuan", "SoHoaDonLienQuan", "NgayHoaDonLienQuan", "NgayHoaDon",
    "MDDKDoanh", "TDDKDoanh", "DCDDKDoanh", "MaKhachHang", "TenDonVi", "NguoiMuaHang",
    "MDVQHNSach", "DiaChiKhachHang", "MaSoThue", "CCCD", "SHChieu", "MailKhachHang",
    "HinhThucThanhToan", "SoDienThoai", "LoaiHangHoa", "LoaiHangHoaDacTrung",
    "MaHangHoa", "TenHangHoa", "SoKhung", "SoMay", "BKSPTVanChuyen", "TNGHang",
    "DCNGuiHang", "MSTNGuiHang", "SDDNguiHang", "GhiChu", "DonViTinh", "SoLuong",
    "DonGia", "ThanhTien", "ThueSuat", "TienThue", "TienBangChu",
]


def _chuoi(value):
    """Đổi giá trị về chuỗi để ghi vào Excel. None -> chuỗi rỗng."""
    if value is None:
        return ""
    return str(value)


def _format_don_gia(value):
    """
    Định dạng Đơn giá khi ghi ra file:
      - Số nguyên -> không có phần thập phân (vd 277778).
      - Có lẻ     -> giữ đúng 2 chữ số thập phân (vd 92592.66).
    (Cần cho tính năng Bảng soát vì Đơn giá HH1 = Thành tiền/Số lượng có thể lẻ.)
    """
    if value is None or str(value).strip() == "":
        return ""
    try:
        f = float(value)
    except (TypeError, ValueError):
        return str(value)
    if f == int(f):
        return str(int(f))
    return f"{f:.2f}"


def tao_cac_dong_output(bang_hoa_don, so_nhom, ngay_hoa_don_str, thue_suat,
                        tron_tong=False, don_gia_chua_thue=False):
    """
    Từ bảng hóa đơn (list dict), tạo ra list các dòng output (mỗi dòng là 1 dict
    theo đúng tên cột OUTPUT_HEADERS).

    Tham số:
      - bang_hoa_don: list dict (mỗi dict là 1 dòng trên bảng hóa đơn)
      - so_nhom: tổng số nhóm hàng hóa hiện có (>= 3)
      - ngay_hoa_don_str: chuỗi 'dd/mm/yyyy' = ngày hôm nay
      - thue_suat: số nguyên (vd 8)
    """
    cac_dong = []
    ma_hd = 0  # sẽ tăng lên 1 cho mỗi dòng hóa đơn CÓ ít nhất 1 nhóm dữ liệu

    for dong in bang_hoa_don:
        # Tìm các nhóm có dữ liệu trong dòng này
        nhom_co_dl = [g for g in range(1, so_nhom + 1) if L.nhom_co_du_lieu(dong, g)]
        if not nhom_co_dl:
            continue  # dòng không có hàng hóa nào -> bỏ qua, không tốn MaHD

        ma_hd += 1

        # Phần thông tin chung (giống nhau cho mọi dòng output cùng MaHD)
        thong_tin_chung = {
            "MaHD": ma_hd,
            "NgayHoaDon": ngay_hoa_don_str,
            "TenDonVi": dong.get("Tên đơn vị mua hàng", ""),
            "NguoiMuaHang": dong.get("Họ tên người mua hàng", ""),
            "DiaChiKhachHang": dong.get("Địa chỉ", ""),
            "MaSoThue": dong.get("Mã số thuế", ""),
            "MailKhachHang": dong.get("Email", ""),  # cột Email (tính năng Bảng soát); ezcloud để rỗng
            "HinhThucThanhToan": dong.get("Hình thức thanh toán", ""),
            "LoaiHangHoa": "1",
            "LoaiHangHoaDacTrung": "0",
            "ThueSuat": thue_suat,
        }

        # Mỗi nhóm có dữ liệu -> 1 dòng output
        for g in nhom_co_dl:
            # "Thành tiền" trên bảng là giá ĐÃ GỒM thuế.
            # ThanhTien trong file output cần là giá CHƯA thuế:
            #   ThanhTien = Thành tiền / (1 + thuế/100), rồi làm tròn gần nhất.
            thanh_tien_gom_thue = L.to_int(dong.get(L.key_nhom(g, "Thành tiền")))
            gia_chua_thue = thanh_tien_gom_thue / (1 + thue_suat / 100)
            thanh_tien = L.lam_tron_thuong(gia_chua_thue)
            if tron_tong:
                # Lấy Tiền thuế = phần còn lại để ThanhTien + TienThue = đúng số tiền
                # gồm thuế ban đầu (tránh lệch 1 đồng kiểu 300001).
                tien_thue = thanh_tien_gom_thue - thanh_tien
            else:
                # TienThue tính trên giá CHƯA thuế (ThanhTien mới), làm tròn lên.
                tien_thue = L.lam_tron_len(thanh_tien * thue_suat / 100)

            dong_out = dict(thong_tin_chung)  # copy phần chung
            dong_out["TenHangHoa"] = dong.get(L.key_nhom(g, "Tên"), "")
            dong_out["DonViTinh"] = dong.get(L.key_nhom(g, "Đơn vị"), "")
            dong_out["SoLuong"] = dong.get(L.key_nhom(g, "Số lượng"), "")
            so_luong = L.to_int(dong.get(L.key_nhom(g, "Số lượng")))
            if don_gia_chua_thue and so_luong != 0:
                # Đơn giá trong file phải là đơn giá CHƯA thuế = ThanhTien (chưa thuế) / Số lượng.
                # (Trên bảng, Đơn giá là giá ĐÃ gồm thuế nên không dùng trực tiếp được.)
                dong_out["DonGia"] = _format_don_gia(thanh_tien / so_luong)
            else:
                dong_out["DonGia"] = _format_don_gia(dong.get(L.key_nhom(g, "Đơn giá"), ""))
            dong_out["ThanhTien"] = thanh_tien
            dong_out["TienThue"] = tien_thue
            cac_dong.append(dong_out)

    return cac_dong


def xuat_file_bytes(bang_hoa_don, so_nhom, ngay_hoa_don_str, thue_suat,
                    tron_tong=False, don_gia_chua_thue=False):
    """
    Tạo file Excel trong bộ nhớ và trả về dạng bytes (để Streamlit cho tải về).
    tron_tong=True: làm cho ThanhTien + TienThue = đúng số tiền gồm thuế (không lệch 1đ).
    don_gia_chua_thue=True: ghi Đơn giá trong file = ThanhTien (chưa thuế) / Số lượng.
    """
    cac_dong = tao_cac_dong_output(
        bang_hoa_don, so_nhom, ngay_hoa_don_str, thue_suat,
        tron_tong=tron_tong, don_gia_chua_thue=don_gia_chua_thue,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    font = Font(name="Arial")

    # Ghi dòng tiêu đề (row 1)
    for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=header)
        c.font = font

    # Ghi dữ liệu từ row 2
    for row_idx, dong_out in enumerate(cac_dong, start=2):
        for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
            gia_tri = _chuoi(dong_out.get(header, ""))
            c = ws.cell(row=row_idx, column=col_idx, value=gia_tri)
            c.font = font
            # Ép kiểu hiển thị là TEXT để các số không bị Excel tự định dạng
            c.number_format = "@"

    # Lưu vào bộ nhớ thay vì ghi ra ổ đĩa
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
