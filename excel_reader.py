"""
excel_reader.py
---------------
Đọc file Excel doanh thu chi tiết từ ezcloud và LỌC ra các phòng đã checkout
trong NGÀY ĐÃ CHỌN.

Đặc điểm file ezcloud:
  - 5 dòng đầu là tiêu đề công ty (bị merge).
  - Dòng header (chứa "Mã đặt phòng", "Tên phòng", ...) thường ở dòng 6, bắt đầu từ cột B.
  - Ngay dưới header có 1 dòng "Phân hệ : Lễ tân" bị merge -> cần bỏ qua.
  - Dữ liệu nằm sau đó. Cuối file có dòng tổng + dòng merge trống.

Cách làm an toàn (không phụ thuộc vị trí cứng):
  - Tự DÒ tìm dòng header bằng cách tìm ô có chữ "Mã đặt phòng".
  - Đọc cột theo TÊN header (không theo số thứ tự cột) -> nếu ezcloud đổi thứ tự cột vẫn chạy đúng.
  - BỎ QUA mọi dòng không có giá trị ở cột "Mã đặt phòng" (tự loại dòng "Phân hệ", dòng tổng, dòng trống).
"""

from datetime import datetime
import openpyxl


# Tên các cột mà chúng ta cần lấy từ file ezcloud.
# Lưu ý: tên phải GIỐNG HỆT header trong file (kể cả dấu).
COT_CAN_LAY = [
    "Mã đặt phòng",
    "Tên phòng",
    "Tên khách",
    "Ngày đến",
    "Ngày đi",
    "Công ty",
    "Số đêm",
    "Giá trung bình",
    "Trạng thái",
]


def parse_datetime_ezcloud(value):
    """
    Chuyển 1 giá trị ngày của ezcloud về kiểu datetime của Python.

    ezcloud điền dạng text "dd/mm/yy hh:mm" (vd "06/06/26 07:56").
    Nhưng đôi khi Excel có thể trả về sẵn kiểu datetime -> ta xử lý cả 2 trường hợp.
    Trả về None nếu không đọc được.
    """
    if value is None:
        return None
    # Trường hợp đã là datetime sẵn
    if isinstance(value, datetime):
        return value
    # Trường hợp là chuỗi text -> thử các định dạng có thể gặp
    text = str(value).strip()
    for fmt in ("%d/%m/%y %H:%M", "%d/%m/%Y %H:%M", "%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _tim_dong_header(ws):
    """
    Dò tìm dòng header: dòng nào có 1 ô chứa đúng chữ "Mã đặt phòng" thì đó là header.
    Trả về (so_dong_header, dict_ten_cot -> so_cot). Nếu không tìm thấy -> (None, None).
    """
    for row in range(1, min(ws.max_row, 30) + 1):  # chỉ cần dò 30 dòng đầu là đủ
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None and str(value).strip() == "Mã đặt phòng":
                # Tìm thấy header. Lập bản đồ tên cột -> số cột cho cả dòng này.
                header_map = {}
                for c in range(1, ws.max_column + 1):
                    v = ws.cell(row=row, column=c).value
                    if v is not None and str(v).strip() != "":
                        header_map[str(v).strip()] = c
                return row, header_map
    return None, None


def doc_va_loc_file(duong_dan_hoac_file, ngay_loc):
    """
    Đọc file ezcloud và lọc ra các booking checkout đúng ngày 'ngay_loc'.

    Tham số:
      - duong_dan_hoac_file: đường dẫn file .xlsx HOẶC đối tượng file (vd từ st.file_uploader).
      - ngay_loc: kiểu datetime.date (chỉ so phần ngày).

    Trả về (danh_sach_booking, canh_bao):
      - danh_sach_booking: list các dict, mỗi dict là 1 phòng đã checkout đúng ngày.
      - canh_bao: list chuỗi cảnh báo (vd thiếu cột) để hiển thị cho user.
    """
    canh_bao = []
    wb = openpyxl.load_workbook(duong_dan_hoac_file, data_only=True)
    ws = wb.active

    so_dong_header, header_map = _tim_dong_header(ws)
    if so_dong_header is None:
        canh_bao.append("Không tìm thấy dòng tiêu đề chứa 'Mã đặt phòng' trong file.")
        return [], canh_bao

    # Kiểm tra xem có thiếu cột nào không
    for ten_cot in COT_CAN_LAY:
        if ten_cot not in header_map:
            canh_bao.append(f"File thiếu cột '{ten_cot}'.")

    ket_qua = []
    # Đọc dữ liệu từ dòng ngay sau header tới hết file
    for row in range(so_dong_header + 1, ws.max_row + 1):
        # Lấy giá trị "Mã đặt phòng"; nếu trống -> bỏ qua dòng này (dòng Phân hệ/tổng/trống)
        col_madp = header_map.get("Mã đặt phòng")
        if col_madp is None:
            break
        ma_dat_phong = ws.cell(row=row, column=col_madp).value
        if ma_dat_phong is None or str(ma_dat_phong).strip() == "":
            continue

        # Hàm nhỏ để lấy 1 ô theo tên cột (trả về None nếu không có cột đó)
        def lay(ten_cot):
            c = header_map.get(ten_cot)
            return ws.cell(row=row, column=c).value if c else None

        trang_thai = lay("Trạng thái")
        ngay_di_raw = lay("Ngày đi")
        ngay_di = parse_datetime_ezcloud(ngay_di_raw)

        # ĐIỀU KIỆN LỌC: Trạng thái = CHECKOUT  VÀ  ngày của "Ngày đi" = ngày đã chọn
        la_checkout = (trang_thai is not None) and (str(trang_thai).strip().upper() == "CHECKOUT")
        dung_ngay = (ngay_di is not None) and (ngay_di.date() == ngay_loc)

        if la_checkout and dung_ngay:
            ket_qua.append(
                {
                    "Mã đặt phòng": str(lay("Mã đặt phòng") or "").strip(),
                    "Tên phòng": str(lay("Tên phòng") or "").strip(),
                    "Tên khách": str(lay("Tên khách") or "").strip(),
                    "Ngày đến": parse_datetime_ezcloud(lay("Ngày đến")),
                    "Ngày đi": ngay_di,
                    "Công ty": str(lay("Công ty") or "").strip(),
                    "Số đêm": lay("Số đêm"),
                    "Giá trung bình": lay("Giá trung bình"),
                }
            )

    return ket_qua, canh_bao
