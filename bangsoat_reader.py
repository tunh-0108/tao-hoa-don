"""
bangsoat_reader.py
------------------
Đọc file Excel "Bảng soát hóa đơn" và đổ dữ liệu ra danh sách dòng để dựng bảng hóa đơn.

Cấu trúc file (theo file mẫu): tiêu đề ở DÒNG 1, dữ liệu từ DÒNG 2.
Các cột (đọc theo TÊN để an toàn nếu thứ tự đổi):
  PHÒNG OUT, TÊN NGƯỜI XHĐ, Hóa đơn chuyên gia, Hóa đơn cơ bản, SỐ ĐÊM,
  Số lượng Dasani, Tên đồ uống khác, Số lượng đồ uống khác, Tiền phòng,
  Tên đơn vị, MST, Địa chỉ, Email

Quy tắc Loại hóa đơn:
  - Cột "Hóa đơn chuyên gia" = 'x'/'X'  -> Lưu trú chuyên gia
  - Cột "Hóa đơn cơ bản"     = 'x'/'X'  -> Lưu trú cơ bản
  - Cả 2 cùng 'x'  -> BÁO LỖI (bỏ qua dòng đó, hiện cảnh báo)
  - Cả 2 đều trống -> cảnh báo, mặc định "Lưu trú cơ bản"
"""

import io

import openpyxl
import invoice_logic as L

# Tên cột trong file Bảng soát (phải khớp tiêu đề ở dòng 1)
COT_PHONG = "PHÒNG OUT"
COT_TEN = "TÊN NGƯỜI XHĐ"
COT_CCCD = "CCCD/PASSPORT"
COT_CHUYEN_GIA = "Hóa đơn chuyên gia"
COT_CO_BAN = "Hóa đơn cơ bản"
COT_SO_DEM = "SỐ ĐÊM"
COT_SL_DASANI = "Số lượng Dasani"
COT_TEN_DU = "Tên đồ uống khác"
COT_SL_DU = "Số lượng đồ uống khác"
COT_TIEN_PHONG = "Tiền phòng"
COT_DON_VI = "Tên đơn vị"
COT_MST = "MST"
COT_DIA_CHI = "Địa chỉ"
COT_EMAIL = "Email"

TAT_CA_COT = [
    COT_PHONG, COT_TEN, COT_CCCD, COT_CHUYEN_GIA, COT_CO_BAN, COT_SO_DEM, COT_SL_DASANI,
    COT_TEN_DU, COT_SL_DU, COT_TIEN_PHONG, COT_DON_VI, COT_MST, COT_DIA_CHI, COT_EMAIL,
]


def _tim_header(ws):
    """Dò dòng header bằng cách tìm ô có chữ 'PHÒNG OUT'. Trả về (số_dòng, map tên->cột)."""
    for row in range(1, min(ws.max_row, 30) + 1):
        for col in range(1, ws.max_column + 1):
            v = ws.cell(row=row, column=col).value
            if v is not None and str(v).strip() == COT_PHONG:
                header_map = {}
                for c in range(1, ws.max_column + 1):
                    hv = ws.cell(row=row, column=c).value
                    if hv is not None and str(hv).strip() != "":
                        header_map[str(hv).strip()] = c
                return row, header_map
    return None, None


def _co_x(value):
    """True nếu ô có dấu 'x' hoặc 'X'."""
    return value is not None and str(value).strip().lower() == "x"


def doc_file_bangsoat(duong_dan_hoac_file):
    """
    Đọc file Bảng soát. Trả về (danh_sach_dong, canh_bao):
      - danh_sach_dong: list dict, mỗi dict có các khóa khớp với tao_dong_tu_bangsoat.
      - canh_bao: list chuỗi cảnh báo/lỗi để hiển thị.
    Import TOÀN BỘ dòng có "PHÒNG OUT" (không lọc gì thêm).
    """
    canh_bao = []
    wb = openpyxl.load_workbook(duong_dan_hoac_file, data_only=True)
    ws = wb.active

    so_dong_header, header_map = _tim_header(ws)
    if so_dong_header is None:
        canh_bao.append("Không tìm thấy dòng tiêu đề chứa 'PHÒNG OUT' trong file.")
        return [], canh_bao

    for ten_cot in TAT_CA_COT:
        if ten_cot not in header_map:
            canh_bao.append(f"File thiếu cột '{ten_cot}'.")

    ket_qua = []
    for row in range(so_dong_header + 1, ws.max_row + 1):
        col_phong = header_map.get(COT_PHONG)
        if col_phong is None:
            break
        phong = ws.cell(row=row, column=col_phong).value
        if phong is None or str(phong).strip() == "":
            continue  # bỏ dòng không có PHÒNG OUT

        def lay(ten_cot):
            c = header_map.get(ten_cot)
            return ws.cell(row=row, column=c).value if c else None

        # Xác định loại hóa đơn từ 2 cột đánh dấu
        co_cg = _co_x(lay(COT_CHUYEN_GIA))
        co_cb = _co_x(lay(COT_CO_BAN))
        if co_cg and co_cb:
            canh_bao.append(
                f"Dòng phòng '{str(phong).strip()}' đánh 'x' ở CẢ 'Hóa đơn chuyên gia' "
                f"và 'Hóa đơn cơ bản' -> đã bỏ qua dòng này, vui lòng sửa file."
            )
            continue
        elif co_cg:
            loai = L.LOAI_BS_CHUYEN_GIA
        elif co_cb:
            loai = L.LOAI_BS_CO_BAN
        else:
            canh_bao.append(
                f"Dòng phòng '{str(phong).strip()}' không đánh dấu loại hóa đơn "
                f"-> tạm để mặc định 'Lưu trú cơ bản'."
            )
            loai = L.LOAI_BS_CO_BAN

        ket_qua.append(
            {
                "Phòng": str(lay(COT_PHONG) or "").strip(),
                "Họ tên người mua hàng": str(lay(COT_TEN) or "").strip(),
                "CCCD/PASSPORT": str(lay(COT_CCCD) or "").strip(),
                "Loại hóa đơn": loai,
                "Số đêm": lay(COT_SO_DEM),
                "Số lượng Dasani": lay(COT_SL_DASANI),
                "Tên đồ uống khác": str(lay(COT_TEN_DU) or "").strip(),
                "Số lượng đồ uống khác": lay(COT_SL_DU),
                "Tiền phòng": lay(COT_TIEN_PHONG),
                "Tên đơn vị mua hàng": str(lay(COT_DON_VI) or "").strip(),
                "Mã số thuế": str(lay(COT_MST) or "").strip(),
                "Địa chỉ": str(lay(COT_DIA_CHI) or "").strip(),
                "Email": str(lay(COT_EMAIL) or "").strip(),
            }
        )

    return ket_qua, canh_bao


def tao_file_mau_bytes():
    """
    Tạo file Excel MẪU để user tải về và điền dữ liệu.
    File chỉ gồm DÒNG TIÊU ĐỀ với đúng các cột (và đúng thứ tự) mà app mong đợi,
    bao gồm cả cột 'CCCD/PASSPORT'. Trả về dạng bytes để dùng cho nút tải xuống.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bảng soát"
    ws.append(TAT_CA_COT)  # dòng 1 = tiêu đề

    # Tô đậm + nền nhẹ + căn giữa cho dòng tiêu đề, giãn rộng cột cho dễ điền
    in_dam = Font(bold=True)
    nen = PatternFill("solid", fgColor="D9E1F2")
    can_giua = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, len(TAT_CA_COT) + 1):
        o = ws.cell(row=1, column=col_idx)
        o.font = in_dam
        o.fill = nen
        o.alignment = can_giua
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.freeze_panes = "A2"  # cố định dòng tiêu đề khi cuộn

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
